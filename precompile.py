import glob, os

'''
The following script: 
  I. processes images and videos to fit blocks of content on the website
  II. generates json files cataloging all projects 

'''
root = os.getcwd()

def clear(dir, exempt = []):
    prev_dir = os.getcwd()
    os.chdir(dir)
    for f in glob.glob('*'):
        os.remove(f)
    os.chdir(prev_dir)

# un-processed files
res_img_major = os.path.join(root, 'res', 'img', 'major') # major projects
res_img_minor = os.path.join(root, 'res', 'img', 'minor') # minor projects
res_img_illust = os.path.join(root, 'res', 'img', 'illust') # illustrations
res_deco = os.path.join(root, 'res', 'img', 'deco') # site decorations

# destination for processed files
thumbnail = os.path.join(root, 'preview', 'thumbnail')

# small blocks for minor projects and experiments
block_height_small = 160
# large blocks for major projects
block_height_large = 320

# I. 1. preprocess images
from PIL import Image
# batch resize all image files in the source directory
# and save them in the destination directory as thumbnails
def compress_img(src, dest, max_height):
    prev_dir = os.getcwd()
    os.chdir(src)
    imgs = glob.glob('*')
    
    tosave = []
    processed = []
    for img in imgs:
        print('processing image: ', img)
        file, ext = os.path.splitext(img)
        with Image.open(img) as im:
            im = im.convert('RGB')
            # resize by height
            w, h = im.size
            w = int(max_height / h * w) 
            im.thumbnail((w, h))
            tosave.append((im, file + '.jpeg'))
            processed.append(file)
    
    os.chdir(dest)
    for t in tosave:
        t[0].save(t[1], 'JPEG')
    
    os.chdir(prev_dir)
    return processed

# clear out preexisting thumbnails
clear(thumbnail)
thumbnail_names = []
thumbnail_names += compress_img(res_img_major, thumbnail, block_height_large)
thumbnail_names += compress_img(res_img_minor, thumbnail, block_height_small)

# I. 2. precompress videos
res_vid_major = os.path.join(root, 'res', 'vid', 'major') # major projects
res_vid_minor = os.path.join(root, 'res', 'vid', 'minor') # minor projects

video = os.path.join(root, 'preview', 'video')

import ffmpeg
# batch resize all image files in the source directory
# and save them in the destination directory as thumbnails
def compress_video(src, dest, max_height):
    prev_dir = os.getcwd()
    os.chdir(src)
    vids = glob.glob('*')

    processed = []
    for vid in vids:
        print('processing video: ' + vid)
        file, ext = os.path.splitext(vid)
        input = ffmpeg.input(vid)
        input = input.video
        input = input.filter('fps', 12)
        input = input.filter('scale', -1, max_height)
        input.output(os.path.join(dest, file + '.webm')).run()
        processed.append(file)
    os.chdir(prev_dir)

    return processed

clear(video)
video_names = []
video_names += compress_video(res_vid_major, video, block_height_large)
video_names += compress_video(res_vid_minor, video, block_height_small)

# generate thumbnails for videos that do not have designated thumbnails
print('================================')
print('preprocessing has finished...')
print('the following videos do not have designated thumbnail: ')
for v in video_names:
    if not v in thumbnail_names:
        print('\t', v)


#II. 1. read major projects
from openpyxl import load_workbook, Workbook
import json

project_sheet = os.path.join(root, 'res', 'proj', 'log.xlsx')
project_json = os.path.join(root, 'res', 'proj')

wb = load_workbook(project_sheet)
major_projs = wb['major']
minor_projs = wb['minor']

# assume that row1 are the keys
# and subsequent rows do not widen/narrow
def table_to_json(worksheet, name, dest):
    rows = worksheet.values
    keys = next(rows)
    w = len(keys)
    h = worksheet.max_row
    print(keys)
    projs = []
    for i in range(h-1):
        obj = [worksheet.cell(row = i+2, column = j+1).value for j in range(w)]
        obj = dict(zip(keys, obj))
        projs.append(obj)
    compiled = json.dumps(projs)
    print(compiled)

    prev_dir = os.getcwd()
    os.chdir(dest)
    # save '[name].json' file
    with open(name+'.js', 'w') as f:
        meta = 'var '+name+' = '
        f.write(meta)
        f.write(compiled)
        f.write(';')
    os.chdir(prev_dir)

table_to_json(major_projs, 'major', project_json)
table_to_json(minor_projs, 'minor', project_json)